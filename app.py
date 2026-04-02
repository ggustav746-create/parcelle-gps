import streamlit as st
import pandas as pd
import requests
import time
import io
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Coordonnées GPS Parcelles",
    page_icon="📍",
    layout="wide"
)

st.title("📍 Géolocalisation de Parcelles Cadastrales")
st.markdown("Importez un fichier Excel contenant vos parcelles pour obtenir leurs coordonnées GPS.")

# ─── helpers ────────────────────────────────────────────────────────────────

def parse_parcelle(raw: str) -> dict | None:
    s = str(raw).strip().upper().replace(" ", "")
    if len(s) < 6:
        return None
    m = re.match(r'^(\d{0,4})([A-Z]{1,2})(\d+)$', s)
    if m:
        return {
            "prefix": m.group(1),
            "section": m.group(2),
            "numero": m.group(3).zfill(4),
        }
    return None


@st.cache_data(show_spinner=False)
def get_insee_code(postal_code: str, city: str) -> str | None:
    try:
        r = requests.get("https://geo.api.gouv.fr/communes", params={
            "codePostal": postal_code,
            "nom": city,
            "fields": "code,nom",
            "limit": 5,
        }, timeout=8)
        if r.status_code == 200 and r.json():
            communes = r.json()
            city_upper = city.upper().strip()
            for c in communes:
                if c.get("nom", "").upper().strip() == city_upper:
                    return c["code"]
            return communes[0]["code"]
    except Exception:
        pass
    try:
        r = requests.get("https://geo.api.gouv.fr/communes", params={
            "codePostal": postal_code,
            "fields": "code,nom",
            "limit": 1,
        }, timeout=8)
        if r.status_code == 200 and r.json():
            return r.json()[0]["code"]
    except Exception:
        pass
    return None


def extract_centroid(geom: dict) -> tuple[float, float] | None:
    gtype = geom.get("type")
    coords = geom.get("coordinates", [])
    if not coords:
        return None
    if gtype == "Polygon":
        ring = coords[0]
    elif gtype == "MultiPolygon":
        ring = coords[0][0]
    else:
        return None
    lon = sum(p[0] for p in ring) / len(ring)
    lat = sum(p[1] for p in ring) / len(ring)
    return round(lat, 6), round(lon, 6)


def query_ign(insee: str, section: str, numero: str, retries: int = 3) -> tuple[float, float] | None:
    """Query IGN with retry logic for resilience on large batches."""
    for attempt in range(retries):
        try:
            r = requests.get(
                "https://apicarto.ign.fr/api/cadastre/parcelle",
                params={
                    "code_insee": insee,
                    "section": section,
                    "numero": numero,
                    "_limit": 1,
                },
                timeout=12,
            )
            if r.status_code == 200:
                features = r.json().get("features", [])
                if features:
                    return extract_centroid(features[0].get("geometry", {}))
                return None  # 200 but no features = doesn't exist, no retry needed
            elif r.status_code == 429:
                time.sleep(2 + attempt * 2)
            elif r.status_code >= 500:
                time.sleep(1 + attempt)
            else:
                return None
        except requests.exceptions.Timeout:
            time.sleep(1 + attempt)
        except Exception:
            return None
    return None


def fetch_parcel_coords(postal_code: str, city: str, parcelle: str) -> dict:
    parsed = parse_parcelle(parcelle)
    result = {
        "latitude": None,
        "longitude": None,
        "status": "❌ Non trouvée",
        "source": "",
        "insee": "",
    }

    # ── Step 1: resolve INSEE code ───────────────────────────────────────────
    insee = get_insee_code(postal_code, city)
    result["insee"] = insee or ""

    # ── Step 2: query IGN cadastre ───────────────────────────────────────────
    if parsed and insee:
        section = parsed["section"]
        numero = parsed["numero"]
        prefix = parsed["prefix"]

        # Build INSEE candidates:
        # Primary: resolved INSEE from city/postal
        # Secondary: for communes nouvelles, reconstruct old INSEE from dept + parcel prefix
        # e.g. Orée d'Anjou insee=49301, prefix=069 → try old commune 49069
        insee_candidates = [insee]
        if prefix and len(prefix) == 3:
            dept = insee[:2]
            old_candidate = dept + prefix
            if old_candidate != insee:
                insee_candidates.append(old_candidate)

        # Section variants: single-letter sections sometimes need a leading zero
        sections_to_try = [section]
        if len(section) == 1:
            sections_to_try.append("0" + section)

        for insee_try in insee_candidates:
            for sec in sections_to_try:
                centroid = query_ign(insee_try, sec, numero)
                if centroid:
                    lat, lon = centroid
                    result.update({
                        "latitude": lat,
                        "longitude": lon,
                        "status": "✅ Trouvée",
                        "source": f"IGN Cadastre (INSEE {insee_try}, section {sec})",
                        "insee": insee_try,
                    })
                    return result

    # ── Step 3: fallback — geocode the commune only ──────────────────────────
    try:
        r = requests.get(
            "https://api-adresse.data.gouv.fr/search/",
            params={"q": f"{city} {postal_code}", "limit": 1, "type": "municipality"},
            timeout=8,
        )
        if r.status_code == 200:
            feats = r.json().get("features", [])
            if feats:
                lon, lat = feats[0]["geometry"]["coordinates"]
                result.update({
                    "latitude": round(lat, 6),
                    "longitude": round(lon, 6),
                    "status": "⚠️ Commune seulement",
                    "source": "Géocodage commune",
                })
    except Exception:
        pass

    return result


def build_output_excel(df_result: pd.DataFrame, extra_cols: list) -> bytes:
    import openpyxl
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "Résultats GPS"

    base_headers = ["Code Postal", "Commune", "Parcelle", "Code INSEE",
                    "Latitude", "Longitude", "Statut", "Source", "Lien Google Maps"]
    extra_headers = [c.upper() for c in extra_cols]
    headers = extra_headers + base_headers

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    fill_even = PatternFill("solid", start_color="EBF3FB")
    fill_odd = PatternFill("solid", start_color="FFFFFF")

    for row_idx, row in df_result.iterrows():
        excel_row = row_idx + 2
        fill = fill_even if row_idx % 2 == 0 else fill_odd

        extra_values = [row.get(c, "") for c in extra_cols]
        base_values = [
            row.get("postal_code", ""),
            row.get("city", ""),
            row.get("parcelle", ""),
            row.get("insee", ""),
            row.get("latitude", ""),
            row.get("longitude", ""),
            row.get("status", ""),
            row.get("source", ""),
        ]
        all_values = extra_values + base_values

        for col_idx, val in enumerate(all_values, 1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        lat = row.get("latitude")
        lon = row.get("longitude")
        maps_col = len(headers)
        if lat and lon:
            link_cell = ws.cell(row=excel_row, column=maps_col)
            url = f"https://www.google.com/maps?q={lat},{lon}"
            link_cell.hyperlink = url
            link_cell.value = "📍 Voir sur Maps"
            link_cell.font = Font(color="1155CC", underline="single")
            link_cell.fill = fill
            link_cell.border = border
            link_cell.alignment = Alignment(horizontal="center")

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf.read()


# ─── Session state for resumable processing ──────────────────────────────────

if "results" not in st.session_state:
    st.session_state.results = []
if "processing_done" not in st.session_state:
    st.session_state.processing_done = False
if "extra_cols" not in st.session_state:
    st.session_state.extra_cols = []


# ─── UI ─────────────────────────────────────────────────────────────────────

with st.sidebar:
    st.header("ℹ️ Informations")
    st.markdown("""
**Format du fichier Excel attendu :**

| LOTID | DOSSIERID | postal_code | city | parcelle |
|-------|-----------|-------------|------|----------|
| L001 | D001 | 85320 | LA BRETONNIERE | 000ZE0003 |
| L002 | D002 | 49270 | OREE D ANJOU | 069AO0278 |

Les colonnes **LOTID**, **DOSSIERID** et toute autre colonne supplémentaire sont conservées dans les résultats.

**Sources utilisées :**
- 🗺️ API Cadastre IGN (apicarto.ign.fr)
- 📍 API Adresse data.gouv.fr (fallback commune)

**Statuts possibles :**
- ✅ Trouvée — coordonnées exactes
- ⚠️ Commune seulement — parcelle introuvable
- ❌ Non trouvée — aucune donnée

**Formats de parcelle supportés :**
- `000ZE0003` — préfixe 3 chiffres + section 2 lettres
- `0000E0218` — préfixe 4 chiffres + section 1 lettre
- `069AO0278` — communes nouvelles (ancien INSEE reconstruit)
    """)
    st.divider()
    delay = st.slider("Délai entre requêtes (ms)", 100, 1000, 350, 50,
                      help="Augmentez si vous avez des erreurs sur de grands fichiers")
    batch_size = st.slider("Pause automatique tous les N lots", 50, 200, 100, 10,
                           help="Pause de 3s tous les N parcelles pour éviter les blocages API")

uploaded = st.file_uploader("📂 Importez votre fichier Excel (.xlsx)", type=["xlsx", "xls"])

if uploaded:
    try:
        df = pd.read_excel(uploaded, dtype=str)
        df.columns = [c.strip().lower() for c in df.columns]

        required = {"postal_code", "city", "parcelle"}
        missing = required - set(df.columns)
        if missing:
            st.error(f"Colonnes manquantes dans le fichier : {', '.join(missing)}")
            st.stop()

        # Detect extra columns (everything that is not one of the 3 required)
        extra_cols = [c for c in df.columns if c not in required]
        st.session_state.extra_cols = extra_cols

        df = df.fillna("")
        df = df[df[["postal_code", "city", "parcelle"]].apply(
            lambda r: any(v.strip() for v in r), axis=1
        )].reset_index(drop=True)

        total = len(df)

        st.success(
            f"✅ {total} parcelle(s) chargée(s)"
            + (f" — colonnes supplémentaires : **{', '.join(c.upper() for c in extra_cols)}**" if extra_cols else "")
        )
        st.dataframe(df, use_container_width=True, height=200)

        # ── Button logic: start fresh or resume ─────────────────────────────
        already_done = len(st.session_state.results)
        run_from = None

        if already_done > 0 and already_done < total and not st.session_state.processing_done:
            st.warning(f"⚠️ Traitement interrompu — **{already_done}/{total}** parcelles traitées.")
            col_resume, col_restart = st.columns(2)
            if col_resume.button("▶️ Reprendre depuis où on s'est arrêté", use_container_width=True):
                run_from = already_done
            if col_restart.button("🔄 Recommencer depuis le début", use_container_width=True):
                st.session_state.results = []
                st.session_state.processing_done = False
                run_from = 0
        else:
            if st.button("🚀 Lancer la géolocalisation", type="primary", use_container_width=True):
                st.session_state.results = []
                st.session_state.processing_done = False
                run_from = 0

        # ── Processing loop ──────────────────────────────────────────────────
        if run_from is not None:
            df_to_process = df.iloc[run_from:].reset_index(drop=True)
            progress = st.progress(0, text="Initialisation…")
            status_area = st.empty()
            error_area = st.empty()

            for i, row in df_to_process.iterrows():
                global_i = run_from + i
                postal = str(row["postal_code"]).strip()
                city = str(row["city"]).strip()
                parcelle = str(row["parcelle"]).strip()

                pct = (global_i + 1) / total
                progress.progress(pct, text=f"Traitement {global_i + 1}/{total} — {parcelle}")
                status_area.info(f"🔍 Recherche : **{parcelle}** ({city} {postal})")

                try:
                    coords = fetch_parcel_coords(postal, city, parcelle)
                    st.session_state.results.append({**row.to_dict(), **coords})
                    error_area.empty()
                except Exception as e:
                    error_area.warning(f"⚠️ Erreur sur {parcelle} : {e}")
                    st.session_state.results.append({
                        **row.to_dict(),
                        "latitude": None, "longitude": None,
                        "status": "❌ Erreur", "source": str(e), "insee": "",
                    })

                # Auto-pause every batch_size rows to let APIs breathe
                if (global_i + 1) % batch_size == 0 and (global_i + 1) < total:
                    status_area.info(f"⏸️ Pause automatique après {global_i + 1} parcelles…")
                    time.sleep(3)

                time.sleep(delay / 1000)

            progress.empty()
            status_area.empty()
            st.session_state.processing_done = True
            st.success("🎉 Traitement terminé !")

        # ── Display results ──────────────────────────────────────────────────
        if st.session_state.results:
            df_result = pd.DataFrame(st.session_state.results)
            extra_cols = st.session_state.extra_cols

            found = (df_result["status"].str.startswith("✅")).sum()
            partial = (df_result["status"].str.startswith("⚠️")).sum()
            not_found = len(df_result) - found - partial

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("📦 Total traité", len(df_result))
            col2.metric("✅ Trouvées", found)
            col3.metric("⚠️ Communes seulement", partial)
            col4.metric("❌ Non trouvées", not_found)

            st.subheader("📋 Résultats")
            display_cols = extra_cols + ["postal_code", "city", "parcelle", "insee",
                                         "latitude", "longitude", "status", "source"]
            display_cols = [c for c in display_cols if c in df_result.columns]
            st.dataframe(df_result[display_cols], use_container_width=True)

            map_df = df_result.dropna(subset=["latitude", "longitude"])
            if not map_df.empty:
                st.subheader("🗺️ Aperçu cartographique")
                st.map(map_df.rename(columns={"latitude": "lat", "longitude": "lon"})[["lat", "lon"]])

            st.subheader("⬇️ Télécharger les résultats")
            excel_bytes = build_output_excel(df_result, extra_cols)
            st.download_button(
                label="📥 Télécharger Excel (.xlsx)",
                data=excel_bytes,
                file_name="parcelles_gps.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
            csv_bytes = df_result.to_csv(index=False).encode("utf-8")
            st.download_button(
                label="📥 Télécharger CSV",
                data=csv_bytes,
                file_name="parcelles_gps.csv",
                mime="text/csv",
                use_container_width=True,
            )

            if not st.session_state.processing_done:
                remaining = total - len(st.session_state.results)
                st.warning(f"⚠️ {remaining} parcelle(s) restante(s). Cliquez sur **Reprendre** pour continuer.")

    except Exception as e:
        st.error(f"Erreur lors de la lecture du fichier : {e}")

else:
    st.session_state.results = []
    st.session_state.processing_done = False
    st.info("👆 Commencez par importer votre fichier Excel.")
    with st.expander("📝 Voir un exemple de fichier attendu"):
        sample = pd.DataFrame({
            "LOTID": ["L001", "L002", "L003"],
            "DOSSIERID": ["D001", "D001", "D002"],
            "postal_code": ["85320", "85200", "49270"],
            "city": ["LA BRETONNIERE", "LONGÈVES", "OREE D ANJOU"],
            "parcelle": ["000ZE0003", "0000E0218", "069AO0278"],
        })
        st.dataframe(sample, use_container_width=True)
